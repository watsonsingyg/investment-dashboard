"""Project data governance checks for the weekly report workbook."""

import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from typing import Optional

import openpyxl

try:
    from .dashboard_data import is_week_header, parse_excel, parse_week_start
    from .excel_store import find_excel, first_week_col
    from .statuses import FUNNEL_STAGES, PRIORITY_ORDER, SCORE_DISPLAY_ORDER, score_label
except ImportError:
    from dashboard_data import is_week_header, parse_excel, parse_week_start
    from excel_store import find_excel, first_week_col
    from statuses import FUNNEL_STAGES, PRIORITY_ORDER, SCORE_DISPLAY_ORDER, score_label


ISSUE_DEFS = {
    'missing_owner': ('缺负责人', 'medium'),
    'missing_priority': ('缺优先级', 'medium'),
    'missing_score': ('缺评分', 'medium'),
    'missing_status': ('缺状态', 'high'),
    'stale_project': ('长期无更新', 'medium'),
    'no_weekly_updates': ('无任何周报', 'high'),
    'no_current_week_update': ('无本周更新', 'low'),
    'abnormal_week_header': ('异常周次', 'medium'),
    'duplicate_project': ('疑似重复项目', 'medium'),
}

ISSUE_ORDER = list(ISSUE_DEFS)
SEVERITY_LABELS = {'high': '高', 'medium': '中', 'low': '低'}


def _clean(value) -> str:
    return str(value or '').strip()


def _project_payload(project: Optional[dict]) -> dict:
    project = project or {}
    score = project.get('score', '')
    return {
        'project': project.get('name', ''),
        'status': project.get('status', ''),
        'priority': project.get('priority', ''),
        'owner': project.get('owner', ''),
        'score': score,
        'score_label': project.get('score_label', '') or score_label(score),
        'last_active': project.get('last_active', ''),
    }


def _make_issue(issue_type: str, project: Optional[dict], reason: str, suggestion: str,
                fixable: bool = False, field: str = '', severity: str = '') -> dict:
    label, default_severity = ISSUE_DEFS[issue_type]
    payload = _project_payload(project)
    payload.update({
        'id': '',
        'type': issue_type,
        'type_label': label,
        'severity': severity or default_severity,
        'severity_label': SEVERITY_LABELS.get(severity or default_severity, '中'),
        'reason': reason,
        'suggestion': suggestion,
        'fixable': bool(fixable),
        'field': field,
    })
    return payload


def _override_path(report_dir) -> str:
    return str(find_excel(report_dir).parent / 'logs' / 'governance_overrides.json')


def load_governance_overrides(report_dir) -> dict:
    path = find_excel(report_dir).parent / 'logs' / 'governance_overrides.json'
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding='utf-8'))
        return raw if isinstance(raw, dict) else {}
    except Exception:
        return {}


def save_governance_issue_state(report_dir, issue_key: str, state: str,
                                note: str = '', days: int = 30) -> dict:
    issue_key = str(issue_key or '').strip()
    state = str(state or '').strip()
    if not issue_key:
        raise ValueError('缺少问题标识')
    if state not in ('active', 'ignored', 'confirmed'):
        raise ValueError('问题状态应为 active / ignored / confirmed')

    log_dir = find_excel(report_dir).parent / 'logs'
    log_dir.mkdir(exist_ok=True)
    path = log_dir / 'governance_overrides.json'
    overrides = load_governance_overrides(report_dir)

    if state == 'active':
        overrides.pop(issue_key, None)
    else:
        days = max(1, min(int(days or 30), 365))
        expires_at = ''
        if state == 'ignored':
            expires_at = (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d')
        overrides[issue_key] = {
            'state': state,
            'note': str(note or '').strip(),
            'expires_at': expires_at,
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }

    path.write_text(json.dumps(overrides, ensure_ascii=False, indent=2), encoding='utf-8')
    return {'ok': True, 'issue_key': issue_key, 'state': state, 'path': str(path)}


def _abnormal_week_headers(excel_path) -> list:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for col in range(first_week_col(ws), ws.max_column + 1):
        raw = ws.cell(1, col).value
        label = _clean(raw)
        if label and not is_week_header(label):
            headers.append({'column': col, 'label': label})
    wb.close()
    return headers


def _normalize_name(name: str) -> str:
    value = str(name or '').lower()
    value = re.sub(r'[\s·•\-_—/\\|:：,，.。;；()（）\[\]【】<>《》]+', '', value)
    return value


def _issue_key(issue: dict) -> str:
    subject = _normalize_name(issue.get('project') or 'sheet') or 'sheet'
    reason_hash = hashlib.sha1(str(issue.get('reason', '')).encode('utf-8')).hexdigest()[:10]
    return f'{issue.get("type", "issue")}:{subject}:{reason_hash}'


def _apply_issue_states(report_dir, issues: list[dict]) -> None:
    overrides = load_governance_overrides(report_dir)
    today = datetime.now().date()
    for issue in issues:
        key = _issue_key(issue)
        issue['issue_key'] = key
        entry = overrides.get(key) or {}
        state = entry.get('state', 'active')
        expires_at = entry.get('expires_at', '')
        if state == 'ignored' and expires_at:
            try:
                if datetime.strptime(expires_at, '%Y-%m-%d').date() < today:
                    state = 'active'
            except ValueError:
                state = 'active'
        if state not in ('ignored', 'confirmed'):
            state = 'active'
        issue['issue_state'] = state
        issue['ignored'] = state == 'ignored'
        issue['confirmed'] = state == 'confirmed'
        issue['ignore_note'] = entry.get('note', '') if state != 'active' else ''
        issue['ignore_expires_at'] = expires_at if state == 'ignored' else ''


def _duplicate_candidates(projects: list[dict]) -> dict[str, set[str]]:
    by_key = defaultdict(list)
    for project in projects:
        key = _normalize_name(project.get('name', ''))
        if key:
            by_key[key].append(project.get('name', ''))

    duplicates = defaultdict(set)
    for names in by_key.values():
        if len(names) > 1:
            for name in names:
                duplicates[name].update(n for n in names if n != name)

    normalized = [(p.get('name', ''), _normalize_name(p.get('name', ''))) for p in projects]
    for idx, (left_name, left_key) in enumerate(normalized):
        if len(left_key) < 5:
            continue
        for right_name, right_key in normalized[idx + 1:]:
            if len(right_key) < 5 or left_name == right_name:
                continue
            ratio = SequenceMatcher(None, left_key, right_key).ratio()
            if ratio >= 0.92:
                duplicates[left_name].add(right_name)
                duplicates[right_name].add(left_name)
    return duplicates


def _assign_ids(issues: list[dict]) -> None:
    for idx, issue in enumerate(issues, 1):
        subject = issue.get('project') or issue.get('reason', '')
        slug = re.sub(r'[^a-z0-9]+', '-', _normalize_name(subject))[:32].strip('-') or 'sheet'
        issue['id'] = f'{issue["type"]}-{idx}-{slug}'


def load_governance_report(report_dir, stale_weeks: int = 8) -> dict:
    excel_path = find_excel(report_dir)
    data = parse_excel(str(excel_path))
    projects = data['projects']
    latest_week = data.get('latest_week', '')
    latest_dt = parse_week_start(latest_week) if latest_week else datetime.min
    stale_cutoff = latest_dt - timedelta(days=stale_weeks * 7) if latest_dt != datetime.min else datetime.min

    issues = []
    for project in projects:
        name = project.get('name', '')
        owner = _clean(project.get('owner', ''))
        priority = _clean(project.get('priority', ''))
        raw_status = _clean(project.get('status_raw', ''))
        status = _clean(project.get('status', ''))
        timeline = project.get('timeline') or []
        last_active = _clean(project.get('last_active', ''))

        if not owner:
            issues.append(_make_issue(
                'missing_owner', project,
                f'{name} 当前负责人为空，后续推进和周报补充缺少明确责任人。',
                '补充项目负责人；如项目已暂停，可指定归档责任人后降低跟进频率。',
                fixable=True, field='owner'
            ))

        if not priority or priority not in PRIORITY_ORDER:
            issues.append(_make_issue(
                'missing_priority', project,
                f'{name} 当前优先级为空或不在 高 / 中 / 低 范围内，无法进入统一分层。',
                '按战略协同、推进阶段和交易确定性补充优先级，建议先在高 / 中 / 低中选择一档。',
                fixable=True, field='priority'
            ))

        if not project.get('score'):
            issues.append(_make_issue(
                'missing_score', project,
                f'{name} 当前项目打分为空，无法纳入评分分布和优先项目筛选。',
                f'按当前判断补充项目打分，建议从 {" / ".join(SCORE_DISPLAY_ORDER)} 中选择一档。',
                fixable=True, field='score'
            ))

        if not raw_status or status not in FUNNEL_STAGES:
            issues.append(_make_issue(
                'missing_status', project,
                f'{name} 当前状态为空或不是标准漏斗阶段，漏斗统计会失真。',
                '将状态调整为 前期沟通、正式尽调、协议签署/交割 或 投后 中的一项。',
                fixable=True, field='status'
            ))

        if not timeline:
            issues.append(_make_issue(
                'no_weekly_updates', project,
                f'{name} 没有任何有效周报记录，无法判断项目历史推进节奏。',
                '若项目仍在跟进，补录最近一次关键进展；若已不跟进，明确状态和优先级后降频或归档。',
            ))
        elif latest_week and not any(item.get('week') == latest_week for item in timeline):
            issues.append(_make_issue(
                'no_current_week_update', project,
                f'{name} 在最新周次 {latest_week} 没有更新，本周动作不清晰。',
                '确认是否需要本周补充进展；若无新增动作，在例会中只保留关键事件触发式跟进。',
                severity='low'
            ))

        last_dt = parse_week_start(last_active) if last_active else datetime.min
        if timeline and latest_dt != datetime.min and last_dt != datetime.min and last_dt < stale_cutoff:
            weeks_gap = max((latest_dt - last_dt).days // 7, stale_weeks)
            issues.append(_make_issue(
                'stale_project', project,
                f'{name} 最后活跃周为 {last_active}，距最新周次约 {weeks_gap} 周，超过 {stale_weeks} 周无有效更新。',
                '重新确认项目是否仍需跟进；若继续推进，请补充下一步动作和负责人，若无推进价值则降频。',
            ))

    duplicate_map = _duplicate_candidates(projects)
    by_name = {p.get('name', ''): p for p in projects}
    for name in sorted(duplicate_map):
        peers = sorted(duplicate_map[name])
        issues.append(_make_issue(
            'duplicate_project', by_name.get(name),
            f'{name} 与 {", ".join(peers[:4])} 名称高度相近，可能是重复录入或命名不一致。',
            '人工核对是否为同一项目；如确认为重复，保留一个标准项目名并迁移历史周报记录。',
        ))

    abnormal_headers = _abnormal_week_headers(excel_path)
    for header in abnormal_headers:
        issues.append(_make_issue(
            'abnormal_week_header', None,
            f'第 {header["column"]} 列表头 “{header["label"]}” 不是 YYYY/MM/DD-YYYY/MM/DD 格式，已被排除在周报时间线外。',
            '确认该列是否是备注列；如果是周报周次，请改成标准日期区间格式。',
        ))

    issues.sort(key=lambda item: (
        ISSUE_ORDER.index(item['type']) if item['type'] in ISSUE_ORDER else 99,
        item.get('project') or item.get('reason') or '',
    ))
    _apply_issue_states(report_dir, issues)
    _assign_ids(issues)

    type_counts = Counter(issue['type'] for issue in issues)
    severity_counts = Counter(issue['severity'] for issue in issues)
    active_issues = [issue for issue in issues if issue.get('issue_state') == 'active']
    active_severity_counts = Counter(issue['severity'] for issue in active_issues)
    issue_types = [
        {
            'type': key,
            'label': ISSUE_DEFS[key][0],
            'count': type_counts.get(key, 0),
            'severity': ISSUE_DEFS[key][1],
        }
        for key in ISSUE_ORDER
    ]

    return {
        'meta': {
            'generated': data.get('generated', ''),
            'source_file': data.get('source_file', ''),
            'latest_week': latest_week,
            'oldest_week': data.get('oldest_week', ''),
            'stale_weeks': stale_weeks,
            'stale_cutoff_week': stale_cutoff.strftime('%Y/%m/%d') if stale_cutoff != datetime.min else '',
        },
        'summary': {
            'total_projects': len(projects),
            'total_issues': len(issues),
            'affected_projects': len({issue['project'] for issue in issues if issue.get('project')}),
            'fixable_issues': sum(1 for issue in issues if issue.get('fixable')),
            'high': severity_counts.get('high', 0),
            'medium': severity_counts.get('medium', 0),
            'low': severity_counts.get('low', 0),
            'active_issues': len(active_issues),
            'active_high': active_severity_counts.get('high', 0),
            'active_medium': active_severity_counts.get('medium', 0),
            'active_low': active_severity_counts.get('low', 0),
            'ignored_issues': sum(1 for issue in issues if issue.get('ignored')),
            'confirmed_issues': sum(1 for issue in issues if issue.get('confirmed')),
        },
        'issue_types': issue_types,
        'issues': issues,
    }
