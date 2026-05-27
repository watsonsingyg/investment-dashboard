"""Excel persistence helpers for the weekly report app."""

import fcntl
import os
import re
import shutil
import sys
import time
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path

import openpyxl

_PROJECT_ROOT = str(Path(__file__).parent.parent)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from config import settings  # noqa: E402

try:
    from .statuses import normalize_score, normalize_status, score_label
    from .summary_cache import clean_text
except ImportError:
    from statuses import normalize_score, normalize_status, score_label
    from summary_cache import clean_text


class ExcelLockError(Exception):
    """Excel 文件被其他操作占用时抛出。"""
    pass


@contextmanager
def _lock_excel(report_dir, timeout=10):
    """获取 Excel 文件的排他锁，超时抛出 ExcelLockError。"""
    lock_path = settings.EXCEL_LOCK
    fd = os.open(str(lock_path), os.O_CREAT | os.O_RDWR)
    try:
        start = time.monotonic()
        while True:
            try:
                fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
                break
            except (BlockingIOError, OSError):
                if time.monotonic() - start > timeout:
                    raise ExcelLockError('Excel 文件正被其他操作占用，请稍后重试')
                time.sleep(0.1)
        yield
    finally:
        fcntl.flock(fd, fcntl.LOCK_UN)
        os.close(fd)


SCORE_HEADER = '项目打分'
SCORE_COL = 9
LEGACY_WEEK_START_COL = 9
SCORED_WEEK_START_COL = 10
WEEK_HEADER_RE = re.compile(r'^\d{4}/\d{2}/\d{2}-\d{4}/\d{2}/\d{2}$')


def has_score_column(ws) -> bool:
    return str(ws.cell(1, SCORE_COL).value or '').strip() == SCORE_HEADER


def ensure_score_column(ws) -> None:
    if has_score_column(ws):
        return
    ws.insert_cols(SCORE_COL)
    ws.cell(1, SCORE_COL).value = SCORE_HEADER


def first_week_col(ws) -> int:
    return SCORED_WEEK_START_COL if has_score_column(ws) else LEGACY_WEEK_START_COL


def _week_start(label: str) -> datetime:
    try:
        return datetime.strptime(str(label).split('-', 1)[0].strip(), '%Y/%m/%d')
    except Exception:
        return datetime.min


def _week_headers(ws) -> list[tuple[int, str]]:
    headers = []
    for col in range(first_week_col(ws), ws.max_column + 1):
        label = str(ws.cell(1, col).value or '').strip()
        if WEEK_HEADER_RE.match(label):
            headers.append((col, label))
    return sorted(headers, key=lambda item: _week_start(item[1]), reverse=True)


def find_excel(report_dir) -> Path:
    report_dir = Path(report_dir)
    files = [
        p for p in report_dir.glob('*.xlsx')
        if p.is_file() and 'backups' not in p.parts
    ]
    if not files:
        raise FileNotFoundError(f'在 {report_dir} 中未找到 .xlsx 文件')
    return max(files, key=lambda p: p.stat().st_mtime)


def get_projects(report_dir) -> list:
    wb = openpyxl.load_workbook(find_excel(report_dir), read_only=True, data_only=True)
    ws = wb.active
    names = [
        str(row[2]).strip()
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[2] and str(row[2]).strip()
    ]
    wb.close()
    return sorted(set(names))


def get_project_row(report_dir, project: str) -> dict:
    project = clean_text(project or '')
    path = find_excel(report_dir)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    score_col_exists = has_score_column(ws)
    week_headers = _week_headers(ws)
    for r in range(2, ws.max_row + 1):
        value = ws.cell(r, 3).value
        if value and clean_text(str(value)) == project:
            score = normalize_score(ws.cell(r, SCORE_COL).value) if score_col_exists else ''
            latest_week = ''
            latest_content = ''
            for col, label in week_headers:
                raw_content = ws.cell(r, col).value
                content = str(raw_content).strip() if raw_content is not None else ''
                if content:
                    latest_week = label
                    latest_content = content
                    break
            result = {
                'name': project,
                'row': r,
                'biz_scope': str(ws.cell(r, 4).value or '').strip(),
                'industry': str(ws.cell(r, 5).value or '').strip(),
                'owner': str(ws.cell(r, 6).value or '').strip(),
                'status': normalize_status(str(ws.cell(r, 7).value or '').strip()),
                'priority': str(ws.cell(r, 8).value or '').strip(),
                'score': score,
                'score_label': score_label(score),
                'latest_week': latest_week,
                'latest_content': latest_content,
                'last_active': latest_week,
            }
            wb.close()
            return result
    wb.close()
    raise KeyError(f'未找到项目：{project}')


def get_existing_week_content(report_dir, project: str, week: str) -> dict:
    project = clean_text(project or '')
    path = find_excel(report_dir)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    proj_row = None
    for r in range(2, ws.max_row + 1):
        value = ws.cell(r, 3).value
        if value and clean_text(str(value)) == project:
            proj_row = r
            break

    week_col = None
    for c in range(first_week_col(ws), ws.max_column + 1):
        hdr = ws.cell(1, c).value
        if hdr and str(hdr).strip() == week:
            week_col = c
            break

    content = ''
    if proj_row and week_col:
        value = ws.cell(proj_row, week_col).value
        content = str(value).strip() if value is not None else ''
    wb.close()
    return {
        'exists': bool(content),
        'row': proj_row,
        'col': week_col,
        'content': content,
    }


def update_project_fields(report_dir, project: str, fields: dict) -> dict:
    allowed_keys = {'biz_scope', 'industry', 'owner', 'status', 'priority', 'score'}
    project = clean_text(project or '')
    clean_fields = {k: str(v or '').strip() for k, v in fields.items() if k in allowed_keys}
    if not clean_fields:
        return {'ok': True, 'project': project, 'changes': [], 'backup_path': ''}

    with _lock_excel(report_dir):
        backup_path = backup_workbook(report_dir)
        path = find_excel(report_dir)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        proj_row = None
        for r in range(2, ws.max_row + 1):
            value = ws.cell(r, 3).value
            if value and clean_text(str(value)) == project:
                proj_row = r
                break
        if proj_row is None:
            wb.close()
            raise KeyError(f'未找到项目：{project}')

        if clean_fields.get('score') or (has_score_column(ws) and 'score' in clean_fields):
            ensure_score_column(ws)

        allowed = {
            'biz_scope': 4,
            'industry': 5,
            'owner': 6,
            'status': 7,
            'priority': 8,
        }
        if has_score_column(ws):
            allowed['score'] = SCORE_COL

        changes = []
        for field, col in allowed.items():
            if field not in clean_fields:
                continue
            old = str(ws.cell(proj_row, col).value or '').strip()
            new = normalize_score(clean_fields[field]) if field == 'score' else clean_fields[field]
            old = normalize_score(old) if field == 'score' else old
            if old != new:
                ws.cell(proj_row, col).value = new
                changes.append({'field': field, 'old': old, 'new': new})
        wb.save(path)
        wb.close()
    return {
        'ok': True,
        'project': project,
        'row': proj_row,
        'changes': changes,
        'backup_path': str(backup_path),
    }


def backup_workbook(report_dir, max_backups: int = None) -> Path:
    if max_backups is None:
        max_backups = settings.MAX_BACKUPS
    src = find_excel(report_dir)
    backup_dir = settings.BACKUP_DIR
    backup_dir.mkdir(exist_ok=True)
    stem = src.stem
    suffix = src.suffix
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = backup_dir / f'{stem}__{ts}{suffix}'
    shutil.copy2(src, dst)

    # 轮转清理：保留最近 max_backups 个备份，删除更旧的
    backups = sorted(
        backup_dir.glob(f'{stem}__*{suffix}'),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for old in backups[max_backups:]:
        old.unlink()

    return dst


def push_to_excel(report_dir, project: str, week: str, content: str,
                  stage: str = '', category: str = '', biz_scope: str = '',
                  score: str = '') -> dict:
    """写入周报内容，同时更新项目状态/类别。新项目自动追加行。
    列映射（1-indexed）：3=项目名, 4=业务范畴, 5=细分行业, 7=项目状态, 8=优先级, 9=项目打分, 10+=周报列
    """
    project = clean_text(project or '')
    with _lock_excel(report_dir):
        backup_path = backup_workbook(report_dir)
        path = find_excel(report_dir)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        score = normalize_score(score)
        if score:
            ensure_score_column(ws)

        proj_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 3).value and clean_text(str(ws.cell(r, 3).value)) == project:
                proj_row = r
                break

        is_new = proj_row is None
        if is_new:
            proj_row = ws.max_row + 1
            ws.cell(proj_row, 3).value = project

        if biz_scope:
            ws.cell(proj_row, 4).value = biz_scope
        if category:
            ws.cell(proj_row, 5).value = category
        if stage:
            ws.cell(proj_row, 7).value = stage
        old_score = normalize_score(ws.cell(proj_row, SCORE_COL).value) if has_score_column(ws) else ''
        if score:
            ws.cell(proj_row, SCORE_COL).value = score

        week_col = None
        week_start = first_week_col(ws)
        for c in range(week_start, ws.max_column + 1):
            hdr = ws.cell(1, c).value
            if hdr and str(hdr).strip() == week:
                week_col = c
                break
        if week_col is None:
            ws.insert_cols(week_start)
            ws.cell(1, week_start).value = week
            week_col = week_start

        old_content = ''
        if week_col:
            current = ws.cell(proj_row, week_col).value
            old_content = str(current).strip() if current is not None else ''

        if content:
            ws.cell(proj_row, week_col).value = content
        wb.save(path)
        wb.close()
    return {
        'ok': True,
        'row': proj_row,
        'col': week_col,
        'is_new': is_new,
        'backup_path': str(backup_path),
        'old_content': old_content,
        'new_content': content,
        'old_score': old_score,
        'new_score': score,
    }


def update_week_content(report_dir, project: str, week: str, content: str) -> dict:
    """更新指定项目、指定周次的周报内容（仅覆盖已存在的单元格）。

    与 push_to_excel() 不同，此函数仅更新已有内容，不会新建项目行或周次列。
    """
    project = clean_text(project or '')
    week = (week or '').strip()
    content = (content or '').strip()

    with _lock_excel(report_dir):
        backup_path = backup_workbook(report_dir)
        path = find_excel(report_dir)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # 查找项目行
        proj_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 3).value and clean_text(str(ws.cell(r, 3).value)) == project:
                proj_row = r
                break
        if proj_row is None:
            wb.close()
            raise KeyError(f'未找到项目：{project}')

        # 查找周次列
        week_col = None
        week_start = first_week_col(ws)
        for c in range(week_start, ws.max_column + 1):
            hdr = ws.cell(1, c).value
            if hdr and str(hdr).strip() == week:
                week_col = c
                break
        if week_col is None:
            wb.close()
            raise KeyError(f'未找到周次列：{week}')

        # 读取旧值
        old_content = str(ws.cell(proj_row, week_col).value or '').strip()

        # 写入新值
        ws.cell(proj_row, week_col).value = content
        wb.save(path)
        wb.close()

    return {
        'ok': True,
        'project': project,
        'week': week,
        'row': proj_row,
        'col': week_col,
        'old_content': old_content,
        'new_content': content,
        'backup_path': str(backup_path),
    }
