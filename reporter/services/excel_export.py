"""
reporter/services/excel_export.py — Excel 导出辅助。

将租户数据导出为 Excel 文件（用于备份或离线查看）。
自动限定当前租户。
"""

from pathlib import Path
from typing import Optional
from flask import g
from sqlalchemy.orm import Session
from models.project import Project
from models.weekly_entry import WeeklyEntry


def _tenant_id() -> Optional[int]:
    """获取当前请求的 tenant_id。"""
    return getattr(g, "tenant_id", None)


def export_to_excel(db: Session, output_path: str) -> dict:
    """
    将当前租户的所有项目导出为 Excel 文件。

    Args:
        db: 数据库会话
        output_path: 输出 Excel 路径

    Returns:
        {path, projects, entries}
    """
    from openpyxl import Workbook

    tid = _tenant_id()
    query = db.query(Project).filter(Project.deleted_at.is_(None))
    if tid is not None:
        query = query.filter(Project.tenant_id == tid)
    projects = query.order_by(Project.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline"

    # 表头
    headers = [
        "项目编号", "时间", "项目", "业务范畴", "细分行业",
        "负责人", "项目状态", "优先级", "项目打分",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 收集所有周次并排序
    all_weeks = set()
    project_entries = {}
    for p in projects:
        entries = (
            db.query(WeeklyEntry)
            .filter(WeeklyEntry.project_id == p.id)
            .order_by(WeeklyEntry.week)
            .all()
        )
        project_entries[p.id] = {e.week: e.content or "" for e in entries}
        all_weeks.update(project_entries[p.id].keys())

    sorted_weeks = sorted(all_weeks)
    for i, w in enumerate(sorted_weeks):
        ws.cell(row=1, column=10 + i, value=w)

    # 数据行
    entry_count = 0
    for row_idx, p in enumerate(projects, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=3, value=p.name)
        ws.cell(row=row_idx, column=4, value=p.biz_scope)
        ws.cell(row=row_idx, column=5, value=p.industry)
        ws.cell(row=row_idx, column=6, value=p.owner)
        ws.cell(row=row_idx, column=7, value=p.status)
        ws.cell(row=row_idx, column=8, value=p.priority)
        ws.cell(row=row_idx, column=9, value=p.score)

        entries = project_entries.get(p.id, {})
        for i, w in enumerate(sorted_weeks):
            content = entries.get(w, "")
            if content:
                ws.cell(row=row_idx, column=10 + i, value=content)
                entry_count += 1

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    return {
        "path": str(out),
        "projects": len(projects),
        "entries": entry_count,
    }
