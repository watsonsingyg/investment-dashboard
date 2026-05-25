"""Excel parsing and dashboard data preparation."""

import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    from .excel_store import SCORE_HEADER, find_excel as find_excel_path
    from .statuses import normalize_score, normalize_status, score_label
    from .summary_cache import clean_text
except ImportError:
    from excel_store import SCORE_HEADER, find_excel as find_excel_path
    from statuses import normalize_score, normalize_status, score_label
    from summary_cache import clean_text


WEEK_HEADER_RE = re.compile(r'^\d{4}/\d{2}/\d{2}-\d{4}/\d{2}/\d{2}$')


def is_week_header(header) -> bool:
    label = str(header or '').strip()
    if not WEEK_HEADER_RE.match(label):
        return False
    try:
        start, end = label.split('-', 1)
        datetime.strptime(start, '%Y/%m/%d')
        datetime.strptime(end, '%Y/%m/%d')
        return True
    except ValueError:
        return False


def parse_week_start(header: str) -> datetime:
    try:
        start_str = str(header).split('-')[0].strip()
        return datetime.strptime(start_str, '%Y/%m/%d')
    except Exception:
        return datetime.min


def truncate(s: str, n: int) -> str:
    s = clean_text(s)
    return s[:n] + '…' if len(s) > n else s


def esc(s) -> str:
    return (str(s)
            .replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;').replace('"', '&quot;'))


def find_excel(script_dir: str) -> str:
    return str(find_excel_path(script_dir))


def week_iso(label: str) -> str:
    try:
        parts = str(label).replace('-', '/').split('/')
        return f'{parts[0]}-{parts[1]}-{parts[2]}'
    except Exception:
        return '0000-00-00'


def week_short(label: str) -> str:
    try:
        parts = str(label).split('/')
        return f'{parts[1]}/{parts[2].split("-")[0]}'
    except Exception:
        return label[:5]


_CATEGORY_RULES = [
    ('AI 语音 / 对话',    ['外呼', '语音', '质检', 'cc saa', ' cc ', 'cc saas', '对话', 'ai 客服', 'ai客服', '口语陪练', '语音模型']),
    ('AI Agent / 大模型', ['agent', '大模型', 'infra', '多模态', 'llm', '落地服务']),
    ('AI infra',          ['数据中台', '中台', '数据平台', '数据湖', '数据仓', 'data infra']),
    ('AI 营销 / GEO',     ['geo', 'seo', '营销', '建站', 'kol']),
    ('AI 金融 / 投研',    ['投顾', '投行', 'ai fa', 'ai+行业', 'ai 行业']),
    ('AI 招聘 / HR',      ['招聘', '猎头']),
    ('AI 工具 / 效率',    ['coding', 'rpa', '低代码', 'chatbi', 'cad', '法律', '空间智能', '留服', '合规诊断', 'aigc', '检测平台', '智能体']),
    ('工业 AI',           ['工业', '智能制造']),
    ('CRM / 客服 / BPO',  ['crm', 'scrm', '呼叫中心', 'bpo', '客服系统', '贷后', '营销】']),
    ('私募 / CVC',        ['私募', 'cvc', '基金']),
    ('不良资产 / 牌照',   ['amc', '不良', '融资担保', '融担', '破产']),
    ('数据 / 另类资产',   ['数据资产', '另类数据', '专家访谈', '专利', '隐私计算', '数字化', '数据合规']),
    ('门店 / 零售',       ['门店', '零售', '视觉检测']),
    ('保险 / 场景金融',   ['保险', '权益', '计费', '场景保险']),
]


def classify_industry(ind: str) -> str:
    s = ind.lower()
    for category, keywords in _CATEGORY_RULES:
        if any(k in s for k in keywords):
            return category
    return '其他'


def parse_excel(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    max_col = ws.max_column
    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
    score_col_idx = 8 if len(headers) > 8 and str(headers[8] or '').strip() == SCORE_HEADER else None
    week_start_idx = 9 if score_col_idx is not None else 8

    week_cols = []
    for i, h in enumerate(headers):
        if i >= week_start_idx and is_week_header(h):
            week_cols.append((i, str(h).strip()))
    week_cols.sort(key=lambda x: parse_week_start(x[1]), reverse=True)

    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[2]
        if not name or not str(name).strip():
            continue

        biz_scope = clean_text(row[3] or '')
        industry = clean_text(row[4] or '')
        owner = clean_text(row[5] or '')
        raw_status = clean_text(row[6] or '')
        status = normalize_status(raw_status)
        priority = clean_text(row[7] or '')
        score = normalize_score(row[score_col_idx]) if score_col_idx is not None and score_col_idx < len(row) else ''
        is_ai = '创新' in biz_scope

        weekly = {}
        for ci, lbl in week_cols:
            if ci < len(row) and row[ci]:
                weekly[lbl] = clean_text(str(row[ci]))

        latest_content, latest_label = '', ''
        if week_cols:
            ci, lbl = week_cols[0]
            if ci < len(row) and row[ci]:
                latest_content = clean_text(str(row[ci]))
                latest_label = lbl

        last_active = ''
        for ci, lbl in week_cols:
            if ci < len(row) and row[ci]:
                last_active = lbl
                break

        first_active = ''
        for ci, lbl in reversed(week_cols):
            if ci < len(row) and row[ci]:
                first_active = lbl
                break

        is_new = False
        if first_active and week_cols:
            newest_dt = parse_week_start(week_cols[0][1])
            first_dt = parse_week_start(first_active)
            is_new = (newest_dt - first_dt).days <= 56

        last_content = ''
        if not latest_content:
            for ci, lbl in week_cols:
                if ci < len(row) and row[ci]:
                    last_content = clean_text(str(row[ci]))
                    break

        timeline = [
            {'week': lbl, 'content': content}
            for lbl, content in sorted(
                weekly.items(),
                key=lambda x: parse_week_start(x[0]),
                reverse=True
            )
        ]

        projects.append({
            'name': clean_text(name),
            'biz_scope': biz_scope,
            'industry': industry,
            'category': classify_industry(industry),
            'owner': owner,
            'status': status,
            'status_raw': raw_status,
            'priority': priority,
            'score': score,
            'score_label': score_label(score),
            'is_ai': is_ai,
            'weekly': weekly,
            'latest_content': latest_content,
            'latest_label': latest_label,
            'last_content': last_content,
            'last_active': last_active,
            'last_active_iso': week_iso(last_active) if last_active else '0000-00-00',
            'first_active': first_active,
            'is_active': bool(latest_content),
            'is_new': is_new,
            'timeline': timeline,
        })

    wb.close()
    return {
        'projects': projects,
        'week_cols': week_cols,
        'latest_week': week_cols[0][1] if week_cols else '',
        'oldest_week': week_cols[-1][1] if week_cols else '',
        'generated': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'source_file': os.path.basename(path),
    }
