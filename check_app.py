#!/usr/bin/env python3
"""Local health checks for the investment weekly-report app."""

import shutil
import sys
from collections import Counter
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

import reporter.app as app_module
from reporter.app import app
from reporter.dashboard_data import parse_excel
from reporter.excel_store import SCORE_HEADER, find_excel, get_project_row, push_to_excel, update_project_fields
from reporter.governance import load_governance_report
from reporter.shadow_store import counts
from reporter.statuses import normalize_status


def check(condition, message):
    if not condition:
        raise AssertionError(message)
    print('OK', message)


def main():
    excel = find_excel(ROOT)
    data = parse_excel(str(excel))
    check(len(data['projects']) == 73, '项目数为 73')
    check(data['latest_week'] == '2026/05/25-2026/05/29', '最新周正确')
    check(data['oldest_week'] != '2/28 本周工作要点', '脏周列表头已排除')
    check(data['oldest_week'] == '2023/10/16-2023/10/20', '最早周为有效日期区间')
    governance = load_governance_report(ROOT)
    issue_types = {issue['type'] for issue in governance['issues']}
    check('missing_priority' in issue_types, '健康检查包含缺优先级')
    check('stale_project' in issue_types, '健康检查包含长期无更新')
    check('no_current_week_update' in issue_types, '健康检查包含无本周更新')
    check('abnormal_week_header' in issue_types, '健康检查标记异常周次')

    statuses = Counter(p['status'] for p in data['projects'])
    for value in ('前期沟通', '正式尽调', '协议签署/交割', '投后'):
        check(value in statuses, f'状态归一包含 {value}')
    check(normalize_status('财务/法律尽调') == '正式尽调', '历史尽调状态归一')
    check(normalize_status('投决通过&协议沟通') == '协议签署/交割', '历史协议状态归一')

    with TemporaryDirectory() as td:
        tmp = Path(td)
        shutil.copy2(excel, tmp / excel.name)
        result = push_to_excel(
            tmp, '健康检查测试项目', '2026/05/25-2026/05/29',
            '仅用于 check_app.py 验证', '前期沟通', '测试行业', '境内/创新业务', '5'
        )
        backups = list((tmp / 'backups').glob('*.xlsx'))
        check(result['ok'], '临时 Excel 写入成功')
        check(len(backups) == 1, '写入前生成备份')
        wb = openpyxl.load_workbook(tmp / excel.name, read_only=True, data_only=True)
        ws = wb.active
        check(ws.cell(1, 9).value == SCORE_HEADER, '首次评分写入插入项目打分列')
        found = False
        score_ok = False
        week_col_ok = ws.cell(1, 10).value == '2026/05/25-2026/05/29'
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == '健康检查测试项目':
                found = True
                score_ok = str(row[8]) == '5'
                break
        wb.close()
        check(found, '临时 Excel 新项目写入成功')
        check(score_ok, '临时 Excel 项目评分写入成功')
        check(week_col_ok, '评分列后周报从第 10 列开始')

    with TemporaryDirectory() as td:
        tmp = Path(td)
        shutil.copy2(excel, tmp / excel.name)
        old_report_dir = app_module.REPORT_DIR
        app_module.REPORT_DIR = tmp
        try:
            client = app.test_client()
            with client.session_transaction() as sess:
                sess['authed'] = True
            missing_score = client.post('/api/push', json={
                'project': '悦点科技',
                'week': '2026/05/25-2026/05/29',
                'content': '仅用于 check_app.py 验证',
                'stage': '前期沟通',
            })
            check(missing_score.status_code == 400, '推送 API 缺评分会拒绝')
            pushed = client.post('/api/push', json={
                'project': '悦点科技',
                'week': '2026/05/25-2026/05/29',
                'content': '仅用于 check_app.py 验证',
                'stage': '前期沟通',
                'score': '4',
            })
            check(pushed.status_code == 200, '推送 API 可写入合法评分')
            check(get_project_row(tmp, '悦点科技')['score'] == '4', '推送 API 写入项目评分')
            patch = client.patch('/api/projects/悦点科技', json={'priority': '中', 'owner': '健康检查'})
            check(patch.status_code == 200, '项目编辑 API 可在临时 Excel 上更新')
            score_patch = client.patch('/api/projects/悦点科技', json={'score': '5'})
            check(score_patch.status_code == 200, '项目编辑 API 可更新评分')
            check(get_project_row(tmp, '悦点科技')['score'] == '5', '项目编辑评分写入 Excel')
            check((tmp / 'backups').exists(), '项目编辑生成临时备份')
            field_options = client.get('/api/field-options')
            check(field_options.status_code == 200, '字段选项 API 可访问')
            field_payload = field_options.get_json()
            check(field_payload['scores'][0]['value'] == '5' and field_payload['scores'][-1]['value'] == '1', '字段选项评分按高到低排序')
            project_detail = client.get('/api/projects/悦点科技')
            check(project_detail.status_code == 200, '项目详情 API 可访问')
            detail_payload = project_detail.get_json()
            check('latest_week' in detail_payload and 'latest_content' in detail_payload, '项目详情包含最近周报')
            gov_before = client.get('/api/governance-data').get_json()
            score_issues = [issue for issue in gov_before['issues'] if issue['type'] == 'missing_score']
            check(score_issues and 'issue_key' in score_issues[0], '健康检查缺评分问题包含稳定标识')
            state_resp = client.post('/api/governance/issue-state', json={
                'issue_key': score_issues[0]['issue_key'],
                'state': 'ignored',
                'note': 'check_app 验证',
            })
            check(state_resp.status_code == 200, '健康检查问题可忽略')
            gov_ignored = client.get('/api/governance-data').get_json()
            ignored = [issue for issue in gov_ignored['issues'] if issue['issue_key'] == score_issues[0]['issue_key']][0]
            check(ignored['ignored'] and ignored['ignore_note'] == 'check_app 验证', '健康检查问题忽略状态可读取')
            restore_resp = client.post('/api/governance/issue-state', json={
                'issue_key': score_issues[0]['issue_key'],
                'state': 'active',
            })
            check(restore_resp.status_code == 200, '健康检查问题可恢复')
            audit = client.get('/api/audit/project/悦点科技')
            check(audit.status_code == 200, '项目审计 API 可访问')
            audit_payload = audit.get_json()
            check(len(audit_payload['diffs']) >= 1, '项目审计包含字段变更')
            check(len(audit_payload['operations']) >= 1, '项目审计包含操作日志')
        finally:
            app_module.REPORT_DIR = old_report_dir

    client = app.test_client()
    check(client.get('/login').status_code == 200, '登录页可访问')
    launch_path = ROOT.parent / '战投工作台.app' / 'Contents' / 'MacOS' / 'launch'
    start_script = (ROOT / 'start.sh').read_text(encoding='utf-8')
    check('kill -9' not in start_script, 'start.sh 不强制杀端口进程')
    if launch_path.exists():
        launch_script = launch_path.read_text(encoding='utf-8')
        check('bash start.sh --background --open' in launch_script, '桌面 app 调用统一启动入口')
        check('kill -9' not in launch_script, '桌面 app 不强制杀端口进程')
    health = client.get('/api/health')
    check(health.status_code == 200, '健康检查 API 可访问')
    health_payload = health.get_json()
    check(health_payload['app'] == 'zhangtou-workbench', '健康检查标识工作台服务')
    with client.session_transaction() as sess:
        sess['authed'] = True
    for path in (
        '/', '/dashboard', '/reporter', '/api/projects',
        '/governance', '/api/governance-data', '/status',
        '/api/field-options', '/api/server-log',
        '/static/vendor/chart.umd.js', '/static/dashboard.js',
        '/static/dashboard.css', '/static/reporter.js', '/static/reporter.css',
        '/static/governance.js', '/static/governance.css',
    ):
        check(client.get(path).status_code == 200, f'{path} 可访问')
    status = client.get('/api/system-status')
    check(status.status_code == 200, '运行状态 API 可访问')
    status_payload = status.get_json()
    check(status_payload['model'] == app_module.GENERATION_MODEL, '运行状态模型名正确')
    check(status_payload['excel_file'].endswith('.xlsx'), '运行状态包含 Excel 文件')
    portal_html = client.get('/').get_data(as_text=True)
    check('href="/dashboard" class="portal-card" target="_blank" rel="noopener"' in portal_html, '工作台看板入口新页面打开')
    check('href="/reporter" class="portal-card" target="_blank" rel="noopener"' in portal_html, '工作台生成器入口新页面打开')
    check('href="/governance" class="portal-card" target="_blank" rel="noopener"' in portal_html, '工作台健康检查入口新页面打开')
    check('href="/status" class="portal-card" target="_blank" rel="noopener"' in portal_html, '工作台运行状态入口新页面打开')

    unauth = app.test_client()
    check(unauth.get('/api/server-log').status_code == 401, '服务日志 API 未登录会拒绝')
    check(unauth.get('/status').status_code == 302, '运行状态页未登录会跳转登录')

    dash = client.get('/api/dashboard-data')
    check(dash.status_code == 200, '/api/dashboard-data 可访问')
    payload = dash.get_json()
    check(payload['metrics']['total'] == 73, 'Dashboard API 项目数正确')
    check('scored_count' in payload['metrics'] and 'avg_score' in payload['metrics'], 'Dashboard API 包含评分指标')
    check(set(payload['counts']['scores']) == {'1', '2', '3', '4', '5'}, 'Dashboard API 包含评分分布')
    check(payload['filters']['scores'] == ['5 夯爆了', '4 顶级', '3 人上人', '2 NPC', '1 拉完了'], '评分展示筛选按高到低排序')
    check('score' in payload['projects'][0] and 'score_label' in payload['projects'][0], 'Dashboard API 项目包含评分字段')
    db_counts = counts(ROOT)
    check(db_counts['projects'] == len(payload['projects']), 'SQLite projects 数量一致')
    check(db_counts['weekly_entries'] == sum(len(p['timeline']) for p in payload['projects']), 'SQLite weekly_entries 数量一致')

    detail = client.get('/api/projects/' + payload['projects'][0]['name'])
    check(detail.status_code == 200, '项目详情 API 可访问')
    detail_payload = detail.get_json()
    check('latest_week' in detail_payload and 'latest_content' in detail_payload, '项目详情包含最近周报字段')
    audit = client.get('/api/audit/project/' + payload['projects'][0]['name'])
    check(audit.status_code == 200, '项目审计 API 可访问')
    recent_audit = client.get('/api/audit/recent')
    check(recent_audit.status_code == 200, '最近审计 API 可访问')
    bad_patch = client.patch('/api/projects/' + payload['projects'][0]['name'], json={'status': '错误状态'})
    check(bad_patch.status_code == 400, '项目编辑 API 拒绝非法状态')
    gov = client.get('/api/governance-data')
    check(gov.status_code == 200, '健康检查数据 API 可访问')
    gov_payload = gov.get_json()
    gov_types = {issue['type'] for issue in gov_payload['issues']}
    for issue_type in ('missing_priority', 'missing_score', 'stale_project', 'no_current_week_update'):
        check(issue_type in gov_types, f'健康检查 API 包含 {issue_type}')
    check(all('issue_key' in issue for issue in gov_payload['issues']), '健康检查问题包含稳定 issue_key')

    csv_resp = client.get('/api/export/projects.csv')
    check(csv_resp.status_code == 200 and 'text/csv' in csv_resp.content_type, '项目 CSV 导出可用')
    md_resp = client.get('/api/export/project/' + payload['projects'][0]['name'] + '.md')
    check(md_resp.status_code == 200 and 'text/markdown' in md_resp.content_type, '项目 Markdown 导出可用')
    weekly_resp = client.get('/api/export/weekly.md?week=' + payload['meta']['latest_week'])
    check(weekly_resp.status_code == 200 and 'text/markdown' in weekly_resp.content_type, '周动态 Markdown 导出可用')

    print('全部检查通过')


if __name__ == '__main__':
    main()
